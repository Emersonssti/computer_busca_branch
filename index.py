from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import threading
import tkinter as tk
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    from PIL import Image, ImageTk

    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False


def _kwargs_subprocess_sem_console() -> dict:
    """Evita janelas cmd piscando no .exe PyInstaller (console=False)."""
    if sys.platform != "win32":
        return {}
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        return {"creationflags": subprocess.CREATE_NO_WINDOW}
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = 0
    return {"startupinfo": si}


def _caminho_git_no_repo(root: str, arquivo_alvo: str) -> str:
    """Caminho relativo à raiz do repositório (formato que o Git espera)."""
    arquivo_alvo = arquivo_alvo.replace("\\", "/").strip()
    if not root:
        return arquivo_alvo
    root_norm = os.path.normpath(root)
    if os.path.isabs(arquivo_alvo):
        alvo_norm = os.path.normpath(arquivo_alvo)
        try:
            rel = os.path.relpath(alvo_norm, root_norm)
        except ValueError:
            return arquivo_alvo.replace("\\", "/")
        if rel.startswith(".."):
            return arquivo_alvo.replace("\\", "/")
        return rel.replace("\\", "/")
    return arquivo_alvo


def _intervalo_ultimos_dias(ndias: int) -> tuple[datetime, datetime]:
    """since = hoje - ndias 00:00; until exclusivo = amanhã 00:00 (inclui commits de hoje)."""
    today = date.today()
    since_dt = datetime.combine(today - timedelta(days=ndias), datetime.min.time())
    until_dt = datetime.combine(today + timedelta(days=1), datetime.min.time())
    return since_dt, until_dt


def _mm_yyyy_exibicao_padrao_60_dias() -> tuple[str, str]:
    """Texto inicial dos campos De/até: mês do (hoje − 60 dias) e mês de hoje."""
    today = date.today()
    ini = today - timedelta(days=60)
    return f"{ini.month}/{ini.year}", f"{today.month}/{today.year}"


def _normaliza_mm_yyyy_digitado(s: str) -> str | None:
    """Retorna 'm/aaaa' canônico ou None se inválido."""
    m = re.match(r"^\s*(\d{1,2})/(\d{4})\s*$", s.strip())
    if not m:
        return None
    mes, ano = int(m.group(1)), int(m.group(2))
    if not 1 <= mes <= 12:
        return None
    return f"{mes}/{ano}"


def _git_run(repo_path: str, *args: str) -> str:
    p = subprocess.run(
        ["git", "-C", repo_path, *args],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode != 0:
        msg = (p.stderr or p.stdout or "").strip() or f"git falhou (código {p.returncode})"
        raise RuntimeError(msg)
    return p.stdout


def _fmt_git_date(dt: datetime) -> str:
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.isoformat(sep=" ", timespec="seconds")


def _parse_git_commit_date(s: str) -> datetime:
    s = s.strip()
    if not s:
        raise ValueError("data vazia")
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S %z", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"data git não reconhecida: {s!r}")


def _refs_remotas_origin(repo_path: str) -> set[str]:
    try:
        out = _git_run(
            repo_path,
            "for-each-ref",
            "--format=%(refname:short)",
            "refs/remotes/origin",
        )
    except RuntimeError:
        return set()
    names = {ln.strip() for ln in out.splitlines() if ln.strip()}
    return {n for n in names if n.rsplit("/", 1)[-1] != "HEAD"}


def _resolver_caminho_tracked_casefold(repo_path: str, caminho: str) -> str:
    """
    Resolve o caminho exatamente como está no índice Git, sem diferenciar maiúsculas/minúsculas.
    Aceita caminho relativo completo ou só o nome do arquivo (se for único no repo).
    """
    caminho = caminho.replace("\\", "/").strip()
    if not caminho:
        raise ValueError("Caminho do arquivo vazio.")

    tracked = [ln.strip() for ln in _git_run(repo_path, "ls-files").splitlines() if ln.strip()]
    if caminho in tracked:
        return caminho

    fold_map = {t.casefold(): t for t in tracked}
    if caminho.casefold() in fold_map:
        return fold_map[caminho.casefold()]

    base = caminho.rsplit("/", 1)[-1].casefold()
    hits = [t for t in tracked if t.rsplit("/", 1)[-1].casefold() == base]
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1:
        amostra = "\n".join(hits[:15])
        mais = f"\n… e mais {len(hits) - 15}." if len(hits) > 15 else ""
        raise ValueError(
            f"Há {len(hits)} arquivos com esse nome (ignorando maiúsculas). "
            f"Informe o caminho completo no repositório, por exemplo:\n{amostra}{mais}"
        )

    raise ValueError(
        f"Arquivo não encontrado no repositório (nem por caminho nem por nome): {caminho!r}"
    )


def _parse_refs_integracao(texto: str) -> list[str]:
    """Lista refs pedidas (ex.: main → origin/main). Vírgula separa várias."""
    out: list[str] = []
    for part in texto.split(","):
        p = part.strip()
        if not p:
            continue
        if "/" not in p:
            p = f"origin/{p}"
        p = p.replace("\\", "/")
        out.append(p)
    return out


def _casar_ref_no_remoto(nomes_refs: set[str], pedido: str) -> str | None:
    """Encontra o nome exato da ref em origin (Git é sensível a maiúsculas no ref)."""
    pedido = pedido.replace("\\", "/").strip()
    if not pedido:
        return None
    if "/" not in pedido:
        candidato = f"origin/{pedido}"
    else:
        candidato = pedido
    alvo = candidato.casefold()
    for n in nomes_refs:
        if n.casefold() == alvo:
            return n
    return None


def _ref_padrao_origin(repo_path: str) -> str | None:
    """Branch padrão do clone (ex.: origin/integracao) via origin/HEAD."""
    p = subprocess.run(
        ["git", "-C", repo_path, "symbolic-ref", "refs/remotes/origin/HEAD"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode != 0:
        return None
    out = p.stdout.strip()
    prefix = "refs/remotes/"
    if out.startswith(prefix):
        return out[len(prefix) :]
    return None


_FALLBACKS_INTEGRACAO = (
    "integracao",
    "INTEGRACAO",
    "homologacao",
    "main",
    "master",
    "develop",
    "release",
    "production",
)


def _resolver_refs_integracao_efetivas(
    repo_path: str, nomes_refs_origin: set[str], texto: str
) -> list[str]:
    """
    Refs que existem no remoto, respeitando maiúsculas reais do repo.
    Se o campo vier vazio ou nada casar, usa origin/HEAD e nomes comuns (integracao, main…).
    """
    resultado: list[str] = []
    vistos: set[str] = set()
    for pedido in _parse_refs_integracao(texto):
        casado = _casar_ref_no_remoto(nomes_refs_origin, pedido)
        if casado and casado not in vistos:
            vistos.add(casado)
            resultado.append(casado)
    if resultado:
        return resultado

    padrao = _ref_padrao_origin(repo_path)
    if padrao:
        casado = _casar_ref_no_remoto(nomes_refs_origin, padrao)
        if casado and casado.rsplit("/", 1)[-1] != "HEAD":
            return [casado]

    for fb in _FALLBACKS_INTEGRACAO:
        casado = _casar_ref_no_remoto(nomes_refs_origin, fb)
        if casado:
            return [casado]
    return []


_LABEL_SEM_REF_INTEGRACAO = "Sem branch de integração no remoto"


def _commit_ancestral_da_ref(repo_path: str, commit_sha: str, ref: str) -> bool:
    """True se commit_sha está no histórico que leva ao tip de ref (ex.: já integrado em produção)."""
    p = subprocess.run(
        ["git", "-C", repo_path, "merge-base", "--is-ancestor", commit_sha, ref],
        capture_output=True,
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode == 0:
        return True
    if p.returncode == 1:
        return False
    return False


def _commit_data_hora_fmt(repo_path: str, sha: str) -> str | None:
    p = subprocess.run(
        ["git", "-C", repo_path, "show", "-s", "--format=%cI", sha],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode != 0 or not p.stdout.strip():
        return None
    try:
        dt = _parse_git_commit_date(p.stdout.strip())
        return dt.strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        return None


def _data_entrada_commit_na_ref(repo_path: str, commit_sha: str, ref: str) -> str | None:
    """
    Data (commit) em que a alteração passou a integrar o histórico da ref:
    primeiro commit no caminho ancestry-path entre commit_sha e o tip de ref (ex.: merge),
    ou o próprio commit_sha se não houver commits intermediários.
    """
    p = subprocess.run(
        [
            "git",
            "-C",
            repo_path,
            "rev-list",
            "--ancestry-path",
            f"{commit_sha}..{ref}",
            "--reverse",
            ref,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    out = p.stdout.strip() if p.returncode == 0 else ""
    linhas = [ln.strip() for ln in out.splitlines() if ln.strip()]
    alvo_sha = linhas[0] if linhas else commit_sha
    return _commit_data_hora_fmt(repo_path, alvo_sha)


def _ultimo_commit_no_arquivo(
    repo_path: str,
    ref: str,
    caminho: str,
    *,
    since: datetime | None,
    until: datetime | None,
) -> tuple[str, str, datetime] | None:
    """Último commit em ref que toca caminho no intervalo, ou None."""
    cmd = ["git", "-C", repo_path, "log", "-1", "--format=%H%n%an%n%cI"]
    if since is not None:
        cmd.append(f"--since={_fmt_git_date(since)}")
    if until is not None:
        cmd.append(f"--until={_fmt_git_date(until)}")
    cmd.append(ref)
    cmd.extend(["--", caminho])
    p = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode != 0 or not p.stdout.strip():
        return None
    parts = p.stdout.strip().split("\n", 2)
    if len(parts) < 3:
        return None
    sha, autor, raw_dt = parts[0], parts[1], parts[2]
    try:
        dt = _parse_git_commit_date(raw_dt)
    except (ValueError, TypeError):
        return None
    return sha, autor, dt


def parse_intervalo_mes_ano(inicio_mm_yyyy: str, fim_mm_yyyy: str) -> tuple[datetime, datetime]:
    """
    Interpreta mm/aaaa (intervalo inclusivo nos meses escolhidos).

    Retorna (since, until) para o Git: since = 1º dia do mês inicial 00:00;
    until = instante exclusivo = 1º dia do mês *seguinte* ao mês final 00:00
    (o Git usa --until como limite não inclusivo).
    """
    pat = re.compile(r"^\s*(\d{1,2})/(\d{4})\s*$")

    def um(s: str) -> tuple[int, int]:
        m = pat.match(s)
        if not m:
            raise ValueError(f"Formato inválido (use mm/aaaa): {s!r}")
        mes, ano = int(m.group(1)), int(m.group(2))
        if not 1 <= mes <= 12:
            raise ValueError(f"Mês inválido: {mes}")
        return mes, ano

    mi, yi = um(inicio_mm_yyyy)
    mf, yf = um(fim_mm_yyyy)
    inicio = datetime(yi, mi, 1, 0, 0, 0)
    if mf == 12:
        until_exclusivo = datetime(yf + 1, 1, 1, 0, 0, 0)
    else:
        until_exclusivo = datetime(yf, mf + 1, 1, 0, 0, 0)
    # Compara com o último momento do mês final (só para validar ordem)
    ultimo_dia = monthrange(yf, mf)[1]
    fim_inclusivo_visual = datetime(yf, mf, ultimo_dia, 23, 59, 59)
    if inicio > fim_inclusivo_visual:
        raise ValueError("Data inicial não pode ser depois da data final.")
    return inicio, until_exclusivo


def buscar_branches_com_alteracao(
    repo_path: str,
    arquivo_alvo: str,
    *,
    since: datetime | None = None,
    until: datetime | None = None,
    refs_integracao: str = "",
    fazer_fetch: bool = True,
    progress_callback=None,
) -> list:
    """
    Branches em origin com pelo menos um commit que toca o arquivo.
    Se since/until forem informados, só commits nesse intervalo
    (since inclusivo; until = primeiro instante após o período, como o Git espera).

    refs_integracao: refs separadas por vírgula; o commit listado é considerado integrado
    se for ancestral do tip de alguma dessas refs (git merge-base --is-ancestor).
    """
    if not os.path.exists(repo_path):
        raise FileNotFoundError("Caminho do repositório inválido.")

    caminho = _caminho_git_no_repo(repo_path, arquivo_alvo)
    caminho = _resolver_caminho_tracked_casefold(repo_path, caminho)
    if progress_callback:
        progress_callback(f"Arquivo resolvido no repo: {caminho}")

    rem_out = [
        ln.strip() for ln in _git_run(repo_path, "remote").splitlines() if ln.strip()
    ]
    if "origin" not in rem_out:
        raise RuntimeError("Nenhum remote 'origin' configurado.")

    if fazer_fetch:
        if progress_callback:
            progress_callback("Atualizando referências remotas (fetch)...")
        try:
            _git_run(repo_path, "fetch", "origin")
        except RuntimeError as e:
            msg = str(e).strip()
            tipo = _classificar_erro_git(msg)
            if tipo == "rede":
                msg = _mensagem_erro_rede(msg)
            elif tipo == "auth":
                msg = _mensagem_erro_auth(msg)
            raise RuntimeError(f"Falha no fetch: {msg}") from e
    elif progress_callback:
        progress_callback("Usando branches já baixadas no clone (sem fetch).")

    nomes_refs_origin = _refs_remotas_origin(repo_path)
    refs_int_ok = _resolver_refs_integracao_efetivas(
        repo_path, nomes_refs_origin, refs_integracao or ""
    )
    if progress_callback:
        if not refs_int_ok:
            progress_callback(
                f"Aviso: {_LABEL_SEM_REF_INTEGRACAO}; colunas de integração não poderão comparar."
            )
        else:
            progress_callback(f"Comparando integração com: {', '.join(refs_int_ok)}")

    branches_com_modificacao = []
    refs = sorted(nomes_refs_origin)
    total = len(refs)

    for i, ref in enumerate(refs):
        if progress_callback:
            progress_callback(f"Analisando branches… {i + 1}/{total}")

        hit = _ultimo_commit_no_arquivo(
            repo_path, ref, caminho, since=since, until=until
        )
        if hit:
            sha, autor, dt = hit
            branches_com_modificacao.append(
                {
                    "branch": ref,
                    "ultimo_commit": sha[:7],
                    "data": dt.strftime("%Y-%m-%d %H:%M"),
                    "autor": autor,
                    "_ts": dt,
                    "_sha": sha,
                }
            )

    def _chave_utc(dt: datetime) -> float:
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc).timestamp()
        return dt.astimezone(timezone.utc).timestamp()

    branches_com_modificacao.sort(key=lambda r: _chave_utc(r["_ts"]), reverse=True)
    for item in branches_com_modificacao:
        del item["_ts"]

    for item in branches_com_modificacao:
        sha = item.pop("_sha")
        if refs_int_ok:
            ref_match = None
            for rref in refs_int_ok:
                if _commit_ancestral_da_ref(repo_path, sha, rref):
                    ref_match = rref
                    break
            if ref_match:
                curto = ref_match.split("/", 1)[-1] if "/" in ref_match else ref_match
                item["integracao"] = f"Sim ({curto})"
                item["data_integracao"] = (
                    _data_entrada_commit_na_ref(repo_path, sha, ref_match) or "—"
                )
            else:
                item["integracao"] = "Não"
                item["data_integracao"] = "—"
        else:
            item["integracao"] = _LABEL_SEM_REF_INTEGRACAO
            item["data_integracao"] = "—"

    return branches_com_modificacao


_APP_VERSION = "v4"
_APP_TITLE = f"Busca Branch Publicada ({_APP_VERSION})"
_COMPANY_NAME = "Computer Intelligence"
_LOGO_FILE = "logo.jpeg"

_GRID_COLS = ("branch", "commit", "data", "autor", "integracao", "data_integracao")
_GRID_HEADERS = (
    "Branch publicada",
    "Commit",
    "Data do commit",
    "Autor",
    "Na integração",
    "Data na integração",
)

_CLR_BG = "#eef1f6"
_CLR_SURFACE = "#ffffff"
_CLR_PRIMARY = "#1565c0"
_CLR_PRIMARY_ACTIVE = "#0d47a1"
_CLR_TEXT = "#1e293b"
_CLR_MUTED = "#64748b"
_CLR_BORDER = "#cbd5e1"
_CLR_ACCENT = "#0891b2"
_CLR_ROW_ALT = "#f1f5f9"
_CLR_STATUS = "#e2e8f0"
_CLR_TREE_SEL = "#bbdefb"

_SETTINGS_DIR = Path(os.environ.get("APPDATA", Path.home())) / "BuscaBranchPublicada"
_SETTINGS_FILE = _SETTINGS_DIR / "settings.json"


def _carregar_settings() -> dict:
    if not _SETTINGS_FILE.is_file():
        return {}
    try:
        with _SETTINGS_FILE.open(encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def _salvar_settings(data: dict) -> None:
    _SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
    with _SETTINGS_FILE.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def _chave_repo(repo_path: str) -> str:
    return os.path.normcase(os.path.normpath(repo_path.strip()))


def _repo_ja_teve_fetch_ok(repo_path: str) -> bool:
    repos = _carregar_settings().get("repos_fetch_ok", {})
    return bool(repos.get(_chave_repo(repo_path)))


def _marcar_fetch_ok(repo_path: str) -> None:
    data = _carregar_settings()
    repos = data.setdefault("repos_fetch_ok", {})
    repos[_chave_repo(repo_path)] = True
    _salvar_settings(data)


def _repo_auth_ok(repo_path: str) -> bool:
    repos = _carregar_settings().get("repos_auth_ok", {})
    return bool(repos.get(_chave_repo(repo_path)))


def _marcar_auth_ok(repo_path: str) -> None:
    data = _carregar_settings()
    repos = data.setdefault("repos_auth_ok", {})
    repos[_chave_repo(repo_path)] = True
    _salvar_settings(data)


def _salvar_pasta_clone(repo_path: str) -> None:
    p = repo_path.strip()
    if not p or not os.path.isdir(p):
        return
    data = _carregar_settings()
    data["ultima_pasta_clone"] = os.path.normpath(p)
    _salvar_settings(data)


def _obter_pasta_clone_inicial() -> str:
    salva = (_carregar_settings().get("ultima_pasta_clone") or "").strip()
    if salva and os.path.isdir(salva) and _eh_repo_git(salva):
        return salva
    return os.path.join(os.path.expanduser("~"), "Documents")


_GIT_TIMEOUT_REDE = 90


def _eh_repo_git(repo_path: str) -> bool:
    git_dir = Path(repo_path) / ".git"
    return git_dir.is_dir() or git_dir.is_file()


def _classificar_erro_git(msg: str) -> str:
    t = msg.lower()
    if any(
        x in t
        for x in (
            "could not resolve host",
            "getaddrinfo",
            "failed to connect",
            "connection timed out",
            "network is unreachable",
            "unable to access",
        )
    ) and "authentication" not in t and "401" not in t and "403" not in t:
        if "could not resolve host" in t or "getaddrinfo" in t:
            return "rede"
    if "could not resolve host" in t or "getaddrinfo failed" in t:
        return "rede"
    if any(
        x in t
        for x in (
            "authentication failed",
            "auth",
            "401",
            "403",
            "credential",
            "denied",
            "cancel",
            "password",
            "terminal prompts disabled",
        )
    ):
        return "auth"
    if "timed out" in t or "failed to connect" in t:
        return "rede"
    return "git"


def _git_comando(repo_path: str | None, *args: str, timeout: int = _GIT_TIMEOUT_REDE) -> subprocess.CompletedProcess:
    cmd = ["git"]
    if repo_path:
        cmd.extend(["-C", repo_path])
    cmd.extend(args)
    env = os.environ.copy()
    env["GIT_TERMINAL_PROMPT"] = "1"
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
        env=env,
        **_kwargs_subprocess_sem_console(),
    )


def _testar_conexao_remota(repo_path: str) -> tuple[str, str]:
    """
    Verifica acesso ao origin.
    Retorna: ok | rede | auth | sem_origin | invalido, mensagem.
    """
    path = repo_path.strip()
    if not path or not os.path.isdir(path):
        return "invalido", "Pasta do clone inválida ou inexistente."
    if not _eh_repo_git(path):
        return "invalido", "A pasta não é um repositório Git (falta .git)."
    rem = _git_comando(path, "remote", timeout=20)
    if rem.returncode != 0 or "origin" not in (rem.stdout or ""):
        return "sem_origin", "Remote 'origin' não configurado neste clone."
    proc = _git_comando(path, "ls-remote", "--heads", "origin", timeout=_GIT_TIMEOUT_REDE)
    if proc.returncode == 0:
        return "ok", "Conectado ao servidor remoto."
    msg = (proc.stderr or proc.stdout or "").strip() or "Falha ao contactar origin."
    return _classificar_erro_git(msg), msg


def _autenticar_no_servidor(repo_path: str) -> tuple[bool, str]:
    """Abre fluxo de credencial do Git (fetch) e retorna sucesso."""
    proc = _git_comando(repo_path, "fetch", "origin", timeout=_GIT_TIMEOUT_REDE)
    if proc.returncode == 0:
        return True, "Autenticação concluída."
    msg = (proc.stderr or proc.stdout or "").strip() or "Fetch cancelado ou falhou."
    return False, msg


def _mensagem_erro_rede(msg: str) -> str:
    return (
        f"{msg}\n\n"
        "Não foi possível alcançar o servidor (DNS/rede).\n"
        "• Verifique conexão com a internet\n"
        "• Conecte a VPN da empresa, se for obrigatória\n"
        "• Tente novamente com “Atualizar remoto (fetch)” desmarcado se já tiver clone local"
    )


def _mensagem_erro_auth(msg: str) -> str:
    return (
        f"{msg}\n\n"
        "É necessário login no Git (conta Microsoft / Azure DevOps da empresa).\n"
        "Conclua a janela de login quando ela abrir — o Windows guarda a credencial "
        "para as próximas vezes."
    )


def _dir_aplicacao() -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
    return Path(__file__).resolve().parent


def _fonte_ui() -> str:
    if sys.platform == "win32":
        return "Segoe UI"
    if sys.platform == "darwin":
        return "Helvetica Neue"
    return "DejaVu Sans"


def _carregar_logo(master: tk.Misc, altura_max: int = 52) -> tk.PhotoImage | None:
    path = _dir_aplicacao() / _LOGO_FILE
    if not path.is_file():
        return None
    if _HAS_PIL:
        with Image.open(path) as img:
            img = img.convert("RGBA")
            img.thumbnail((280, altura_max), Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(img, master=master)
    try:
        return tk.PhotoImage(file=str(path), master=master)
    except tk.TclError:
        return None


def _aplicar_tema_visual(root: tk.Tk) -> ttk.Style:
    root.configure(bg=_CLR_BG)
    style = ttk.Style(root)
    for tema in ("clam", "alt", "default"):
        try:
            style.theme_use(tema)
            break
        except tk.TclError:
            continue
    fam = _fonte_ui()
    base = (fam, 10)
    base_sm = (fam, 9)
    base_lg = (fam, 13, "bold")

    style.configure(".", background=_CLR_BG, foreground=_CLR_TEXT, font=base)
    style.configure("TFrame", background=_CLR_BG)
    style.configure("Surface.TFrame", background=_CLR_SURFACE)
    style.configure("Header.TFrame", background=_CLR_SURFACE)
    style.configure("TLabel", background=_CLR_BG, foreground=_CLR_TEXT, font=base)
    style.configure(
        "HeaderTitle.TLabel",
        background=_CLR_SURFACE,
        foreground=_CLR_TEXT,
        font=base_lg,
    )
    style.configure(
        "HeaderSub.TLabel",
        background=_CLR_SURFACE,
        foreground=_CLR_MUTED,
        font=base_sm,
    )
    style.configure(
        "Company.TLabel",
        background=_CLR_SURFACE,
        foreground=_CLR_ACCENT,
        font=(fam, 9),
    )
    style.configure(
        "Status.TLabel",
        background=_CLR_STATUS,
        foreground=_CLR_TEXT,
        font=base_sm,
        padding=(10, 6),
    )
    style.configure(
        "TLabelframe",
        background=_CLR_BG,
        foreground=_CLR_PRIMARY,
        font=(fam, 10, "bold"),
    )
    style.configure(
        "TLabelframe.Label",
        background=_CLR_BG,
        foreground=_CLR_PRIMARY,
        font=(fam, 10, "bold"),
    )
    style.configure("TLabelframe", bordercolor=_CLR_BORDER)
    style.configure("TEntry", fieldbackground=_CLR_SURFACE, padding=4)
    style.configure("TButton", padding=(10, 6), font=base)
    style.configure(
        "Primary.TButton",
        background=_CLR_PRIMARY,
        foreground="#ffffff",
        font=(fam, 10, "bold"),
        padding=(16, 9),
        borderwidth=0,
    )
    style.map(
        "Primary.TButton",
        background=[("active", _CLR_PRIMARY_ACTIVE), ("disabled", "#94a3b8")],
        foreground=[("disabled", "#f1f5f9")],
    )
    style.configure(
        "Treeview",
        background=_CLR_SURFACE,
        fieldbackground=_CLR_SURFACE,
        foreground=_CLR_TEXT,
        rowheight=28,
        font=base,
        borderwidth=0,
    )
    style.configure(
        "Treeview.Heading",
        background=_CLR_PRIMARY,
        foreground="#ffffff",
        font=(fam, 10, "bold"),
        relief="flat",
    )
    style.map(
        "Treeview",
        background=[("selected", _CLR_TREE_SEL)],
        foreground=[("selected", _CLR_TEXT)],
    )
    style.configure("Vertical.TScrollbar", background=_CLR_BG, troughcolor=_CLR_BG)
    style.configure("Horizontal.TScrollbar", background=_CLR_BG, troughcolor=_CLR_BG)
    style.configure("TNotebook", background=_CLR_BG, borderwidth=0)
    style.configure("TNotebook.Tab", padding=(14, 8), font=(fam, 10))
    style.map(
        "TNotebook.Tab",
        background=[("selected", _CLR_SURFACE), ("!selected", _CLR_BG)],
        foreground=[("selected", _CLR_PRIMARY), ("!selected", _CLR_MUTED)],
    )
    return style


def _linhas_da_tree(tree: ttk.Treeview) -> list[tuple[str, ...]]:
    return [tuple(tree.item(iid, "values")) for iid in tree.get_children()]


def _exportar_grid_excel(
    parent: tk.Tk,
    tree: ttk.Treeview,
    *,
    repo: str,
    arquivo: str,
) -> None:
    if not _HAS_OPENPYXL:
        messagebox.showerror(
            _APP_TITLE,
            "Não foi possível exportar para Excel.\n\n"
            "Se você está usando o .exe, informe o suporte de TI. "
            "Se está em desenvolvimento: pip install -r requirements.txt",
            parent=parent,
        )
        return
    rows = _linhas_da_tree(tree)
    if not rows:
        messagebox.showinfo(
            _APP_TITLE,
            "Não há resultados na tabela. Execute uma busca antes de exportar.",
            parent=parent,
        )
        return
    nome_padrao = f"busca_branch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    destino = filedialog.asksaveasfilename(
        parent=parent,
        title="Exportar resultados para Excel",
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        initialfile=nome_padrao,
    )
    if not destino:
        return
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        meta = [
            (_APP_TITLE,),
            (f"Desenvolvido por {_COMPANY_NAME}",),
            (f"Repositório: {repo}",),
            (f"Arquivo analisado: {arquivo}",),
            (f"Exportado em: {datetime.now().strftime('%Y-%m-%d %H:%M')}",),
            (),
        ]
        for linha in meta:
            ws.append(linha)
        ws.append(list(_GRID_HEADERS))
        linha_cabecalho = ws.max_row
        for row in rows:
            ws.append(list(row))
        fill = PatternFill("solid", fgColor="1565C0")
        font_hdr = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[linha_cabecalho]:
            cell.fill = fill
            cell.font = font_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, len(_GRID_HEADERS) + 1):
            letra = get_column_letter(col_idx)
            max_len = len(_GRID_HEADERS[col_idx - 1])
            for row_idx in range(linha_cabecalho + 1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 80))
            ws.column_dimensions[letra].width = max(12, min(max_len + 2, 48))
        ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)
        wb.save(destino)
    except OSError as e:
        messagebox.showerror(_APP_TITLE, f"Não foi possível salvar o arquivo:\n{e}", parent=parent)
        return
    messagebox.showinfo(
        _APP_TITLE,
        f"Planilha exportada com sucesso.\n\n{destino}",
        parent=parent,
    )


def _fonte_monoespacada() -> str:
    return "Consolas" if sys.platform == "win32" else "Courier New"


def _obter_diff_commit_arquivo(repo_path: str, sha_ref: str, arquivo_alvo: str) -> str:
    """Diff do arquivo no commit indicado (hash curto ou completo)."""
    sha = _git_run(repo_path, "rev-parse", "--verify", sha_ref).strip()
    caminho = _resolver_caminho_tracked_casefold(
        repo_path, _caminho_git_no_repo(repo_path, arquivo_alvo)
    )
    meta = _git_run(
        repo_path,
        "show",
        "-s",
        "--format=commit %H (%h)%nAutor: %an <%ae>%nData: %ci%nBranch/ref: %D",
        sha,
    ).strip()
    p = subprocess.run(
        ["git", "-C", repo_path, "show", "--no-color", sha, "--", caminho],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        **_kwargs_subprocess_sem_console(),
    )
    if p.returncode != 0:
        msg = (p.stderr or p.stdout or "").strip() or f"git show falhou (código {p.returncode})"
        raise RuntimeError(msg)
    patch = p.stdout
    cabecalho = f"Arquivo: {caminho}\n{meta}\n{'—' * 72}\n\n"
    if not patch.strip():
        return cabecalho + "(Nenhuma alteração de conteúdo neste commit para este arquivo.)"
    return cabecalho + patch


def _mostrar_janela_diff(parent: tk.Tk, titulo: str, conteudo: str) -> None:
    win = tk.Toplevel(parent)
    win.title(titulo)
    win.minsize(680, 420)
    win.geometry("920x580")
    win.transient(parent)
    win.configure(bg=_CLR_BG)

    hdr = ttk.Frame(win, padding=(14, 12, 14, 6))
    hdr.pack(fill=tk.X)
    ttk.Label(hdr, text=titulo, style="HeaderTitle.TLabel", wraplength=860).pack(anchor="w")
    ttk.Label(
        hdr,
        text="Selecione o texto para copiar. Linhas com + adicionaram; com − removeram.",
        style="HeaderSub.TLabel",
    ).pack(anchor="w", pady=(4, 0))

    body = scrolledtext.ScrolledText(
        win,
        wrap=tk.NONE,
        font=(_fonte_monoespacada(), 10),
        bg=_CLR_SURFACE,
        fg=_CLR_TEXT,
        insertbackground=_CLR_PRIMARY,
        relief="flat",
        highlightthickness=1,
        highlightbackground=_CLR_BORDER,
        padx=10,
        pady=8,
    )
    body.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))
    body.insert("1.0", conteudo)
    body.tag_configure("add", foreground="#0d7a3b")
    body.tag_configure("del", foreground="#b91c1c")
    body.tag_configure("hunk", foreground=_CLR_PRIMARY)
    for i, linha in enumerate(conteudo.splitlines(), start=1):
        if linha.startswith("+++") or linha.startswith("---") or linha.startswith("@@"):
            body.tag_add("hunk", f"{i}.0", f"{i}.end")
        elif linha.startswith("+") and not linha.startswith("+++"):
            body.tag_add("add", f"{i}.0", f"{i}.end")
        elif linha.startswith("-") and not linha.startswith("---"):
            body.tag_add("del", f"{i}.0", f"{i}.end")
    body.configure(state="disabled")

    rod = ttk.Frame(win, padding=(14, 0, 14, 12))
    rod.pack(fill=tk.X)

    def copiar_diff():
        win.clipboard_clear()
        win.clipboard_append(conteudo)
        win.update_idletasks()

    ttk.Button(rod, text="Copiar tudo", command=copiar_diff).pack(side=tk.LEFT)
    ttk.Button(rod, text="Fechar", command=win.destroy).pack(side=tk.RIGHT)
    win.bind("<Escape>", lambda _e: win.destroy())


_AJUDA_ABAS: list[tuple[str, list[dict]]] = [
    (
        "Comece aqui",
        [
            {
                "titulo": "Para que serve este programa?",
                "corpo": [
                    "O Busca Branch Publicada ajuda você a descobrir em quais versões publicadas "
                    "(branches remotas) um arquivo do projeto foi alterado, em um período que você escolhe.",
                    "Ele também indica se aquela alteração já chegou na branch de integração/produção "
                    "do seu time — por exemplo, integracao.",
                ],
            },
            {
                "titulo": "Como executar o programa (arquivo .exe)",
                "itens": [
                    "Você recebe o arquivo BuscaBranchGit_v4.exe — não precisa instalar Python nem abrir código-fonte.",
                    "Dê duplo clique no .exe (ou use um atalho na área de trabalho).",
                    "Na primeira execução o Windows pode pedir confirmação — escolha Permitir / Executar.",
                    "O programa abre a janela Busca Branch Publicada; use F1 ou o botão Ajuda se tiver dúvida.",
                ],
            },
            {
                "titulo": "Ao abrir o programa (automático)",
                "itens": [
                    "O programa restaura a última Pasta do clone que você usou.",
                    "Em seguida verifica se o Git consegue falar com o servidor (origin).",
                    "Se precisar de login, pergunta se deseja abrir a janela Microsoft/Azure — "
                    "conclua o login; o Windows guarda para as próximas vezes.",
                    "Se aparecer erro de rede (não achou o servidor), verifique internet ou VPN.",
                ],
            },
            {
                "titulo": "Passo a passo para a primeira busca com sucesso",
                "itens": [
                    "Na máquina deve existir o Git for Windows (peça ao suporte de TI se não tiver).",
                    "Tenha a pasta do projeto já clonada no PC (ex.: pasta ORACLE do repositório).",
                    "Em Pasta do clone, clique Procurar… e selecione a raiz dessa pasta (onde está o projeto Git).",
                    "Em Arquivo, digite o caminho do arquivo (ex.: procedures/meu_arquivo.sql).",
                    "Deixe desmarcado Atualizar remoto (fetch) — assim a busca não pede login de novo.",
                    "Clique Executar busca e aguarde a mensagem de conclusão na barra de status.",
                    "Dê duplo clique em uma linha para ver as alterações no código (diff).",
                ],
                "dica": "A pasta do clone fica salva em %APPDATA%\\BuscaBranchPublicada\\settings.json.",
            },
            {
                "titulo": "Palavras que você vai ver",
                "itens": [
                    "Clone — pasta do projeto baixada no seu computador.",
                    "Branch — linha de desenvolvimento/publicação (ex.: origin/feature_x).",
                    "Commit — um registro de alteração no histórico (cada linha da tabela mostra um).",
                    "Origin — servidor remoto de onde vêm as branches publicadas.",
                    "Fetch — atualizar a lista de branches/commits do servidor para o seu clone.",
                    "Integração — branch onde as mudanças costumam ser reunidas antes de produção.",
                ],
                "dica": "Você não precisa ser especialista em Git: preencha os campos, execute a busca "
                "e use a tabela. Esta ajuda explica cada opção em linguagem simples.",
            },
        ],
    ),
    (
        "Critérios de busca",
        [
            {
                "titulo": "Pasta do clone",
                "corpo": [
                    "É a pasta raiz do repositório no seu PC — aquela que contém a pasta oculta .git "
                    "(você normalmente não vê .git no Explorer, mas ela precisa existir).",
                ],
                "itens": [
                    "Clique em Procurar… e selecione a pasta do projeto (ex.: ORACLE).",
                    "Não use uma subpasta (ex.: procedures) — use a raiz do repositório.",
                    "O programa só lê o que já está no seu computador; não substitui o Git Desktop.",
                ],
                "exemplo": "C:\\Users\\seu_usuario\\…\\ORACLE",
            },
            {
                "titulo": "Arquivo",
                "corpo": [
                    "Informe qual arquivo você quer rastrear. Pode ser o caminho a partir da raiz do "
                    "repositório ou o caminho completo dentro do clone.",
                ],
                "itens": [
                    "Maiúsculas/minúsculas no nome não importam — o programa acha o arquivo real no Git.",
                    "Se existir mais de um arquivo com o mesmo nome em pastas diferentes, informe o "
                    "caminho completo (ex.: procedures/meu_script.sql).",
                ],
                "exemplo": "procedures/TS.ASS_VALIDA_PF_ddl.sql",
            },
            {
                "titulo": "Período (De / até) — formato mm/aaaa",
                "corpo": [
                    "Define em qual intervalo de tempo o programa procura commits que alteraram o arquivo.",
                ],
                "itens": [
                    "Ao abrir o programa, os campos mostram um intervalo sugerido; se você NÃO mudar "
                    "esses valores, a busca usa os últimos 60 dias corridos (do calendário).",
                    "Se você alterar De ou até, passa a valer meses inteiros: do dia 1 do mês inicial "
                    "até o último dia do mês final.",
                    "Se apagar os dois campos, volta ao padrão de 60 dias.",
                ],
                "exemplo": "De: 3/2026  até: 5/2026  → março, abril e maio de 2026 completos.",
            },
            {
                "titulo": "Branch de integração",
                "corpo": [
                    "É a branch que o seu time trata como integração ou produção. O programa compara "
                    "cada resultado com ela para preencher as colunas Na integração e Data na integração.",
                ],
                "itens": [
                    "Valor comum: integracao (o programa busca origin/integracao no remoto).",
                    "Várias branches: separe por vírgula (ex.: integracao, main).",
                    "Em branco: o programa tenta adivinhar (origin/HEAD, integracao, main, etc.).",
                ],
            },
            {
                "titulo": "Atualizar remoto (fetch)",
                "corpo": [
                    "Quando marcado, antes da busca o programa pede ao Git para trazer branches novas "
                    "do servidor (Azure DevOps, GitHub, etc.). Pode abrir login na primeira vez.",
                ],
                "itens": [
                    "Desmarcado (recomendado no dia a dia): usa o que já está no clone — rápido e sem login.",
                    "Marcado: use quando precisa de branches/commits que acabaram de ser publicados.",
                    "Após um fetch bem-sucedido, o programa costuma desmarcar sozinho na próxima vez.",
                ],
                "dica": "Repositórios Azure DevOps pedem conta Microsoft da empresa, não conta GitHub/Google.",
            },
        ],
    ),
    (
        "Tabela de resultados",
        [
            {
                "titulo": "O que cada coluna significa",
                "itens": [
                    "Branch publicada — nome da branch remota onde houve alteração no arquivo.",
                    "Commit — identificador curto da alteração (como um “código” daquele registro).",
                    "Data do commit — quando essa alteração foi feita.",
                    "Autor — quem fez o commit.",
                    "Na integração — Sim (nome da branch): já está no histórico de integração; "
                    "Não: ainda não; ou aviso se a branch de integração não foi encontrada.",
                    "Data na integração — quando a alteração entrou no histórico da branch de integração "
                    "(útil após merges).",
                ],
            },
            {
                "titulo": "Ações na tabela",
                "itens": [
                    "Duplo clique na linha — abre a janela Ver alterações (diff do código).",
                    "Botão Ver alterações — mesmo efeito (com uma linha selecionada).",
                    "Botão Copiar branch — copia o nome da branch para a área de transferência.",
                    "Ctrl+C com foco na tabela — copia o nome da branch da linha selecionada.",
                    "Clique direito — menu com ver alterações, copiar branch e exportar Excel.",
                    "Exportar para Excel — salva uma planilha .xlsx com os dados da tabela (escolha onde guardar).",
                ],
            },
            {
                "titulo": "Ordem das linhas",
                "corpo": [
                    "A tabela lista do commit mais recente para o mais antigo dentro do período. "
                    "Cada linha mostra o último commit que tocou o arquivo naquela branch no intervalo.",
                ],
            },
        ],
    ),
    (
        "Ver alterações (diff)",
        [
            {
                "titulo": "O que é o diff?",
                "corpo": [
                    "Diff é a comparação linha a linha do que mudou no arquivo naquele commit. "
                    "Não é o arquivo inteiro — só o trecho alterado naquela gravação do histórico.",
                    "Pense assim: você escolheu um arquivo e uma linha da tabela; o programa mostra "
                    "“o que foi mexido nesse arquivo naquele momento”.",
                ],
            },
            {
                "titulo": "Como abrir",
                "itens": [
                    "Dê duplo clique em qualquer linha da tabela de resultados, ou",
                    "Selecione a linha e clique no botão Ver alterações, ou",
                    "Clique direito na linha → Ver alterações no código.",
                ],
                "dica": "Execute uma busca antes — a tabela precisa ter pelo menos uma linha.",
            },
            {
                "titulo": "Como ler a janela do diff",
                "itens": [
                    "No topo aparecem: caminho do arquivo, commit completo, autor e data.",
                    "Linhas começando com + (verde) — trechos adicionados ou alterados para esse conteúdo.",
                    "Linhas começando com − (vermelho) — trechos removidos ou versão anterior.",
                    "Linhas com @@ — indicam em qual parte do arquivo a mudança ocorreu.",
                    "Linhas com --- ou +++ — cabeçalho técnico do Git (nome do arquivo antigo/novo).",
                ],
                "corpo": [
                    "Use a barra de rolagem horizontal e vertical se o texto for largo. "
                    "Você pode selecionar e copiar trechos; o botão Copiar tudo envia tudo para a área de transferência.",
                ],
            },
            {
                "titulo": "O que o diff NÃO mostra",
                "itens": [
                    "Outros arquivos alterados no mesmo commit — só o arquivo que você digitou em Arquivo.",
                    "Todos os commits da branch — só o commit da linha clicada.",
                    "Versão “ao vivo” no servidor — usa o histórico já presente no seu clone local.",
                ],
            },
            {
                "titulo": "Se aparecer erro ao abrir o diff",
                "itens": [
                    "Confirme se Pasta do clone e Arquivo estão corretos.",
                    "O commit pode não existir no clone se nunca houve fetch e a branch é muito nova.",
                    "Tente marcar Atualizar remoto (fetch), autenticar uma vez e buscar de novo.",
                ],
            },
        ],
    ),
    (
        "Login e internet",
        [
            {
                "titulo": "Por que pede login Microsoft?",
                "corpo": [
                    "Se o repositório da sua empresa está no Azure DevOps (dev.azure.com), o Git precisa "
                    "da sua conta corporativa Microsoft — não é a mesma coisa que GitHub ou login Google.",
                ],
                "itens": [
                    "Conta GitHub ou login Google do GitHub Desktop não serve para Azure DevOps.",
                    "Quando marcar Atualizar remoto (fetch) e clicar Executar busca, pode abrir uma janela de login.",
                    "Use o e-mail corporativo Microsoft da empresa e conclua o login — não feche antes de terminar.",
                    "Se a janela sumir ou der erro, peça ao TI para autenticar o Git na pasta do clone uma vez.",
                ],
            },
            {
                "titulo": "Como evitar login toda hora",
                "itens": [
                    "Deixe desmarcado Atualizar remoto (fetch) nas buscas do dia a dia.",
                    "Marque fetch só quando precisar de branches publicadas há pouco no servidor.",
                    "Depois de um login bem-sucedido, o Windows costuma lembrar a credencial nas próximas vezes.",
                ],
            },
        ],
    ),
    (
        "Requisitos no PC",
        [
            {
                "titulo": "O que precisa ter instalado",
                "itens": [
                    "Windows com o programa BuscaBranchGit_v4.exe (este arquivo que você executa).",
                    "Git for Windows — download em https://git-scm.com/download/win (instalação padrão).",
                    "Pasta do repositório já clonada no computador (projeto ORACLE ou equivalente).",
                    "Permissão de leitura nessa pasta e, se usar fetch, acesso de rede ao servidor da empresa.",
                ],
                "dica": "O .exe já inclui tudo para interface, logo e exportar Excel. Só o Git fica instalado "
                "separadamente no Windows, porque o programa consulta o histórico pelo comando git.",
            },
            {
                "titulo": "O que NÃO é necessário",
                "itens": [
                    "Instalar Python.",
                    "Rodar pip ou abrir o código-fonte do projeto.",
                    "Ter o repositório computer_busca_branch — só o clone do projeto que você analisa.",
                ],
            },
            {
                "titulo": "Atalhos úteis",
                "itens": [
                    "F1 — abre esta ajuda.",
                    "Esc — fecha a janela de ajuda ou de alterações (diff).",
                ],
            },
        ],
    ),
    (
        "Problemas comuns",
        [
            {
                "titulo": "O programa não abre",
                "itens": [
                    "Clique com o botão direito no .exe → Propriedades → Desbloquear (se existir).",
                    "Antivírus pode bloquear — peça ao TI para liberar BuscaBranchGit_v4.exe.",
                    "Tente executar como administrador apenas se o TI orientar.",
                ],
            },
            {
                "titulo": "Erro ao buscar ou “caminho inválido”",
                "itens": [
                    "Confira se Pasta do clone aponta para a raiz do repositório (não uma subpasta).",
                    "Confira se o caminho em Arquivo está correto (ex.: procedures/arquivo.sql).",
                    "Verifique se o Git está instalado: abra o Prompt de Comando e digite git --version.",
                ],
            },
            {
                "titulo": "Exportar Excel não funciona",
                "itens": [
                    "Execute uma busca antes — a tabela precisa ter linhas.",
                    "Escolha uma pasta onde você tem permissão para salvar o arquivo .xlsx.",
                    "Se aparecer mensagem de erro, informe o texto ao suporte — o .exe já traz a biblioteca necessária.",
                ],
            },
            {
                "titulo": "Sobre esta versão",
                "corpo": [
                    f"Versão {_APP_VERSION} — desenvolvido por {_COMPANY_NAME}. "
                    "O programa não guarda senhas; usa o Git instalado no Windows e as credenciais "
                    "que você autorizar quando o sistema pedir login.",
                ],
            },
        ],
    ),
]


def _configurar_tags_ajuda(txt: tk.Text, fam: str) -> None:
    txt.tag_configure("h1", font=(fam, 15, "bold"), foreground=_CLR_PRIMARY, spacing3=10)
    txt.tag_configure("h2", font=(fam, 12, "bold"), foreground=_CLR_PRIMARY, spacing1=14, spacing3=6)
    txt.tag_configure("body", font=(fam, 10), foreground=_CLR_TEXT, spacing1=4, spacing3=8, lmargin1=8)
    txt.tag_configure(
        "bullet",
        font=(fam, 10),
        foreground=_CLR_TEXT,
        spacing1=2,
        lmargin1=22,
        lmargin2=22,
    )
    txt.tag_configure(
        "dica",
        font=(fam, 10),
        foreground="#0c4a6e",
        background="#e0f2fe",
        spacing1=8,
        spacing3=8,
        lmargin1=12,
        lmargin2=12,
        rmargin=12,
    )
    txt.tag_configure(
        "exemplo",
        font=(_fonte_monoespacada(), 9),
        foreground=_CLR_MUTED,
        background=_CLR_ROW_ALT,
        spacing1=6,
        spacing3=8,
        lmargin1=12,
        lmargin2=12,
    )
    txt.tag_configure("sep", font=(fam, 1), spacing1=4)


def _inserir_secao_ajuda(txt: tk.Text, secao: dict) -> None:
    txt.insert(tk.END, secao["titulo"] + "\n", "h2")
    for par in secao.get("corpo", []):
        txt.insert(tk.END, par + "\n\n", "body")
    for item in secao.get("itens", []):
        txt.insert(tk.END, "•  " + item + "\n", "bullet")
    if secao.get("itens"):
        txt.insert(tk.END, "\n", "body")
    if dica := secao.get("dica"):
        txt.insert(tk.END, "  Dica: " + dica + "\n\n", "dica")
    if ex := secao.get("exemplo"):
        txt.insert(tk.END, "  Exemplo: " + ex + "\n\n", "exemplo")


def _preencher_aba_ajuda(txt: tk.Text, secoes: list[dict], fam: str) -> None:
    _configurar_tags_ajuda(txt, fam)
    txt.configure(state="normal")
    txt.delete("1.0", tk.END)
    for i, secao in enumerate(secoes):
        _inserir_secao_ajuda(txt, secao)
        if i < len(secoes) - 1:
            txt.insert(tk.END, "\n", "sep")
    txt.configure(state="disabled")


def _mostrar_janela_ajuda(parent: tk.Tk) -> None:
    fam = _fonte_ui()
    win = tk.Toplevel(parent)
    win.title(f"Ajuda — {_APP_TITLE}")
    win.minsize(720, 520)
    win.geometry("820x620")
    win.transient(parent)
    win.configure(bg=_CLR_BG)

    topo = tk.Frame(win, bg=_CLR_PRIMARY, height=4)
    topo.pack(fill=tk.X)

    hdr = tk.Frame(win, bg=_CLR_SURFACE, padx=20, pady=16)
    hdr.pack(fill=tk.X)
    ttk.Label(hdr, text=f"Central de ajuda — {_APP_VERSION}", style="HeaderTitle.TLabel").pack(anchor="w")
    ttk.Label(
        hdr,
        text="Guia para quem usa o executável (.exe). Não é necessário Python nem acesso ao código-fonte.",
        style="HeaderSub.TLabel",
    ).pack(anchor="w", pady=(6, 0))

    note_fr = ttk.Frame(win, padding=(16, 12, 16, 8))
    note_fr.pack(fill=tk.BOTH, expand=True)
    notebook = ttk.Notebook(note_fr)
    notebook.pack(fill=tk.BOTH, expand=True)

    for nome_aba, secoes in _AJUDA_ABAS:
        tab = ttk.Frame(notebook, padding=0)
        notebook.add(tab, text=f"  {nome_aba}  ")
        canvas = tk.Canvas(tab, bg=_CLR_SURFACE, highlightthickness=0, borderwidth=0)
        scroll = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=_CLR_SURFACE)
        inner.bind(
            "<Configure>",
            lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")),
        )
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        txt = tk.Text(
            inner,
            wrap=tk.WORD,
            width=88,
            height=32,
            font=(fam, 10),
            bg=_CLR_SURFACE,
            fg=_CLR_TEXT,
            relief="flat",
            padx=16,
            pady=12,
            cursor="arrow",
            borderwidth=0,
            highlightthickness=0,
        )
        txt.pack(fill=tk.BOTH, expand=True)
        _preencher_aba_ajuda(txt, secoes, fam)

        def _on_mousewheel(event, c=canvas):
            c.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _on_mousewheel)

    rod = ttk.Frame(win, padding=(16, 8, 16, 14))
    rod.pack(fill=tk.X)
    ttk.Label(
        rod,
        text=f"Desenvolvido por {_COMPANY_NAME}",
        foreground=_CLR_MUTED,
        font=(fam, 9),
    ).pack(side=tk.LEFT)
    ttk.Button(rod, text="Fechar", command=win.destroy).pack(side=tk.RIGHT)
    win.bind("<Escape>", lambda _e: win.destroy())
    win.focus_set()


def _criar_app():
    default_repo = _obter_pasta_clone_inicial()
    _de_padrao, _ate_padrao = _mm_yyyy_exibicao_padrao_60_dias()
    _de_padrao_key = _normaliza_mm_yyyy_digitado(_de_padrao)
    _ate_padrao_key = _normaliza_mm_yyyy_digitado(_ate_padrao)
    fam = _fonte_ui()

    root = tk.Tk()
    root.title(_APP_TITLE)
    root.minsize(820, 600)
    root.geometry("1140x640")
    _aplicar_tema_visual(root)

    outer = ttk.Frame(root, padding=(0, 0))
    outer.grid(row=0, column=0, sticky="nsew")
    root.rowconfigure(0, weight=1)
    root.columnconfigure(0, weight=1)

    header = ttk.Frame(outer, style="Header.TFrame", padding=(20, 14, 20, 12))
    header.grid(row=0, column=0, sticky="ew")
    header.columnconfigure(1, weight=1)
    outer.columnconfigure(0, weight=1)

    logo_holder = ttk.Frame(header, style="Header.TFrame")
    logo_holder.grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 16))
    logo_img = _carregar_logo(logo_holder)
    if logo_img is not None:
        lbl_logo = ttk.Label(logo_holder, image=logo_img)
        lbl_logo.image = logo_img  # noqa: SLF001 — referência exigida pelo Tk
        lbl_logo.pack()
    else:
        ttk.Label(
            logo_holder,
            text="Computer Intelligence",
            style="Company.TLabel",
        ).pack(anchor="w")

    titulo_fr = ttk.Frame(header, style="Header.TFrame")
    titulo_fr.grid(row=0, column=1, sticky="w")
    ttk.Label(titulo_fr, text=_APP_TITLE, style="HeaderTitle.TLabel").pack(anchor="w")
    ttk.Label(
        titulo_fr,
        text="Consulta branches remotas por alteração em arquivo",
        style="HeaderSub.TLabel",
    ).pack(anchor="w", pady=(2, 0))

    hdr_acoes = ttk.Frame(header, style="Header.TFrame")
    hdr_acoes.grid(row=0, column=2, rowspan=2, sticky="ne")
    btn_ajuda_hdr = ttk.Button(hdr_acoes, text="Ajuda  F1", width=12)

    body = ttk.Frame(outer, padding=(20, 12, 20, 16))
    body.grid(row=1, column=0, sticky="nsew")
    outer.rowconfigure(1, weight=1)
    body.columnconfigure(0, weight=1)

    def mostrar_ajuda():
        _mostrar_janela_ajuda(root)

    btn_ajuda_hdr.configure(command=mostrar_ajuda)
    btn_ajuda_hdr.pack()

    crit = ttk.LabelFrame(body, text="  Critérios de busca  ", padding=(14, 12))
    crit.grid(row=0, column=0, sticky="ew")
    crit.columnconfigure(1, weight=1)

    ttk.Label(crit, text="Pasta do clone").grid(row=0, column=0, sticky="w", pady=5)
    repo_var = tk.StringVar(value=default_repo)
    ttk.Entry(crit, textvariable=repo_var).grid(
        row=0, column=1, sticky="ew", padx=(12, 8), pady=5
    )

    def escolher_pasta():
        p = filedialog.askdirectory(title="Selecionar pasta do repositório Git", parent=root)
        if p:
            repo_var.set(p)
            _salvar_pasta_clone(p)

    ttk.Button(crit, text="Procurar…", command=escolher_pasta).grid(row=0, column=2, pady=5)

    ttk.Label(crit, text="Arquivo").grid(row=1, column=0, sticky="nw", pady=5)
    arquivo_var = tk.StringVar(value="procedures/TS.ASS_VALIDA_PF_ddl.sql")
    ttk.Entry(crit, textvariable=arquivo_var).grid(
        row=1, column=1, columnspan=2, sticky="ew", padx=(12, 0), pady=5
    )

    ttk.Label(crit, text="Período (mm/aaaa)").grid(row=2, column=0, sticky="nw", pady=5)
    periodo_fr = ttk.Frame(crit)
    periodo_fr.grid(row=2, column=1, sticky="w", padx=(12, 0), pady=5)
    inicio_var = tk.StringVar(value=_de_padrao)
    fim_var = tk.StringVar(value=_ate_padrao)
    ttk.Label(periodo_fr, text="De").pack(side=tk.LEFT)
    ttk.Entry(periodo_fr, textvariable=inicio_var, width=11).pack(side=tk.LEFT, padx=(6, 20))
    ttk.Label(periodo_fr, text="até").pack(side=tk.LEFT)
    ttk.Entry(periodo_fr, textvariable=fim_var, width=11).pack(side=tk.LEFT, padx=(6, 0))
    ttk.Label(
        periodo_fr,
        text="  ·  padrão: últimos 60 dias se não alterar",
        foreground=_CLR_MUTED,
    ).pack(side=tk.LEFT, padx=(12, 0))

    ttk.Label(crit, text="Branch de integração").grid(row=2, column=2, sticky="w", padx=(16, 0), pady=5)
    integracao_var = tk.StringVar(value="integracao")
    ttk.Entry(crit, textvariable=integracao_var).grid(
        row=2, column=3, sticky="ew", padx=(8, 0), pady=5
    )
    crit.columnconfigure(3, weight=1)

    fetch_var = tk.BooleanVar(value=not _repo_ja_teve_fetch_ok(default_repo))

    acoes = ttk.Frame(body)
    acoes.grid(row=1, column=0, sticky="ew", pady=(16, 10))
    btn_run = ttk.Button(acoes, text="Executar busca", style="Primary.TButton")
    btn_run.pack(side=tk.LEFT, padx=(0, 12))
    ttk.Checkbutton(
        acoes,
        text="Atualizar remoto (fetch)",
        variable=fetch_var,
    ).pack(side=tk.LEFT)
    status_var = tk.StringVar(value="Pronto para buscar.")
    ttk.Label(acoes, textvariable=status_var, style="HeaderSub.TLabel").pack(
        side=tk.LEFT, padx=(12, 0)
    )

    resultados = ttk.LabelFrame(body, text="  Resultados  ", padding=(12, 10))
    resultados.grid(row=2, column=0, sticky="nsew", pady=(4, 0))
    body.rowconfigure(2, weight=1)
    resultados.columnconfigure(0, weight=1)
    resultados.rowconfigure(1, weight=1)

    tree_frame = ttk.Frame(resultados)
    tree_frame.grid(row=1, column=0, sticky="nsew")
    tree_frame.rowconfigure(0, weight=1)
    tree_frame.columnconfigure(0, weight=1)

    tree = ttk.Treeview(tree_frame, columns=_GRID_COLS, show="headings", height=14)
    for col, heading in zip(_GRID_COLS, _GRID_HEADERS, strict=True):
        tree.heading(col, text=heading)
    tree.column("branch", width=260, minwidth=120)
    tree.column("commit", width=72, minwidth=60)
    tree.column("data", width=120, minwidth=100)
    tree.column("autor", width=140, minwidth=80)
    tree.column("integracao", width=200, minwidth=90)
    tree.column("data_integracao", width=130, minwidth=100)
    tree.tag_configure("odd", background=_CLR_ROW_ALT)
    tree.tag_configure("even", background=_CLR_SURFACE)

    sy = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    sx = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
    tree.grid(row=0, column=0, sticky="nsew")
    sy.grid(row=0, column=1, sticky="ns")
    sx.grid(row=1, column=0, sticky="ew")

    rodape = ttk.Label(
        outer,
        text=f"Desenvolvido por {_COMPANY_NAME}",
        foreground=_CLR_MUTED,
        font=(fam, 8),
    )
    rodape.grid(row=2, column=0, sticky="e", padx=20, pady=(0, 8))

    root.bind("<F1>", lambda _e: mostrar_ajuda())

    def branch_da_selecao():
        sel = tree.selection()
        if not sel:
            return None
        vals = tree.item(sel[0], "values")
        return vals[0] if vals else None

    def linha_da_selecao() -> dict[str, str] | None:
        sel = tree.selection()
        if not sel:
            return None
        vals = tree.item(sel[0], "values")
        if len(vals) < len(_GRID_COLS):
            return None
        return dict(zip(_GRID_COLS, vals, strict=True))

    def ver_alteracoes_selecionada():
        row = linha_da_selecao()
        if not row:
            status_var.set("Selecione uma linha na tabela.")
            return
        path = repo_var.get().strip()
        arq = arquivo_var.get().strip()
        if not path or not arq:
            messagebox.showwarning(
                _APP_TITLE,
                "Informe a pasta do clone e o arquivo nos critérios de busca.",
                parent=root,
            )
            return

        set_busy(True)
        status_var.set("Carregando alterações…")

        def worker():
            err = None
            texto = ""
            try:
                texto = _obter_diff_commit_arquivo(path, row["commit"], arq)
            except Exception as e:
                err = e

            def finish():
                set_busy(False)
                if err:
                    messagebox.showerror(_APP_TITLE, str(err), parent=root)
                    status_var.set("Erro ao carregar alterações.")
                    return
                titulo = f"{row['branch']} @ {row['commit']}"
                _mostrar_janela_diff(root, titulo, texto)
                status_var.set("Alterações exibidas.")

            root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def copiar_branch_selecionada():
        b = branch_da_selecao()
        if not b:
            status_var.set("Selecione uma linha na tabela.")
            return
        root.clipboard_clear()
        root.clipboard_append(b)
        root.update_idletasks()
        curto = b if len(b) <= 72 else b[:69] + "…"
        status_var.set(f"Copiado: {curto}")

    menu_ctx = tk.Menu(root, tearoff=0)
    menu_ctx.add_command(label="Ver alterações no código", command=ver_alteracoes_selecionada)
    menu_ctx.add_command(label="Copiar nome da branch", command=copiar_branch_selecionada)
    menu_ctx.add_separator()
    menu_ctx.add_command(
        label="Exportar para Excel",
        command=lambda: _exportar_grid_excel(
            root,
            tree,
            repo=repo_var.get().strip(),
            arquivo=arquivo_var.get().strip(),
        ),
    )

    def menu_contexto(event):
        iid = tree.identify_row(event.y)
        if not iid:
            return
        tree.selection_set(iid)
        try:
            menu_ctx.tk_popup(event.x_root, event.y_root)
        finally:
            menu_ctx.grab_release()

    def on_ctrl_c(_event):
        copiar_branch_selecionada()
        return "break"

    tree.bind("<Button-3>", menu_contexto)
    tree.bind("<Control-c>", on_ctrl_c)
    tree.bind("<Control-C>", on_ctrl_c)

    def duplo_clique(_event):
        iid = tree.identify_row(_event.y)
        if iid:
            tree.selection_set(iid)
            ver_alteracoes_selecionada()

    tree.bind("<Double-1>", duplo_clique)

    bar_resultados = ttk.Frame(resultados)
    bar_resultados.grid(row=0, column=0, sticky="ew", pady=(0, 10))
    ttk.Button(
        bar_resultados,
        text="Ver alterações",
        command=ver_alteracoes_selecionada,
    ).pack(side=tk.LEFT, padx=(0, 8))
    ttk.Button(
        bar_resultados,
        text="Copiar branch",
        command=copiar_branch_selecionada,
    ).pack(side=tk.LEFT, padx=(0, 8))
    btn_export = ttk.Button(
        bar_resultados,
        text="Exportar para Excel",
        command=lambda: _exportar_grid_excel(
            root,
            tree,
            repo=repo_var.get().strip(),
            arquivo=arquivo_var.get().strip(),
        ),
    )
    btn_export.pack(side=tk.LEFT)
    ttk.Label(
        bar_resultados,
        text="Duplo clique na linha para ver o diff · Ctrl+C copia a branch",
        foreground=_CLR_MUTED,
    ).pack(side=tk.RIGHT)

    def _preencher_tree(res: list) -> None:
        tree.delete(*tree.get_children())
        for idx, r in enumerate(res):
            tag = "odd" if idx % 2 else "even"
            tree.insert(
                "",
                "end",
                tags=(tag,),
                values=(
                    r["branch"],
                    r["ultimo_commit"],
                    r["data"],
                    r["autor"],
                    r["integracao"],
                    r["data_integracao"],
                ),
            )

    def set_busy(busy: bool):
        btn_run.state(["disabled"] if busy else ["!disabled"])
        btn_export.state(["disabled"] if busy else ["!disabled"])

    def on_progress(msg: str):
        root.after(0, lambda m=msg: status_var.set(m))

    def fluxo_login_git(path: str):
        status_var.set("Aguardando login no Git…")

        def worker_login():
            ok, msg = _autenticar_no_servidor(path)

            def done():
                if ok:
                    _marcar_auth_ok(path)
                    _marcar_fetch_ok(path)
                    _salvar_pasta_clone(path)
                    fetch_var.set(False)
                    status_var.set("Login salvo. Pronto para buscar.")
                    messagebox.showinfo(
                        _APP_TITLE,
                        "Login realizado com sucesso.\n\n"
                        "A credencial fica salva no Windows para este repositório.",
                        parent=root,
                    )
                else:
                    tipo = _classificar_erro_git(msg)
                    if tipo == "rede":
                        texto = _mensagem_erro_rede(msg)
                    elif tipo == "auth":
                        texto = _mensagem_erro_auth(msg)
                    else:
                        texto = msg
                    messagebox.showerror(_APP_TITLE, texto, parent=root)
                    status_var.set("Login não concluído.")

            root.after(0, done)

        threading.Thread(target=worker_login, daemon=True).start()

    def verificar_git_ao_abrir():
        path = repo_var.get().strip()
        if not path or not os.path.isdir(path):
            status_var.set("Informe a pasta do clone do repositório.")
            return
        if not _eh_repo_git(path):
            status_var.set("A pasta do clone não é um repositório Git válido.")
            return

        status_var.set("Verificando autenticação Git…")

        def worker():
            status, msg = _testar_conexao_remota(path)

            def finish():
                _salvar_pasta_clone(path)
                if status == "ok":
                    _marcar_auth_ok(path)
                    status_var.set("Pronto — Git conectado ao servidor remoto.")
                    return
                if status == "rede":
                    status_var.set("Sem conexão com o servidor (verifique rede/VPN).")
                    messagebox.showwarning(
                        _APP_TITLE,
                        _mensagem_erro_rede(msg),
                        parent=root,
                    )
                    return
                if status in ("auth", "git"):
                    status_var.set("Login Git necessário.")
                    if messagebox.askyesno(
                        _APP_TITLE,
                        "Para usar o programa com o servidor remoto, é preciso fazer login no Git "
                        "(conta Microsoft / Azure DevOps).\n\n"
                        "Deseja abrir a janela de login agora?\n\n"
                        "A credencial permanece salva para as próximas aberturas do programa.",
                        parent=root,
                    ):
                        fluxo_login_git(path)
                    else:
                        status_var.set(
                            "Login pendente. Busque sem fetch ou autentique depois."
                        )
                    return
                if status == "sem_origin":
                    status_var.set("Clone sem remote origin.")
                    messagebox.showwarning(_APP_TITLE, msg, parent=root)

            root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def executar():
        _preencher_tree([])
        path = repo_var.get().strip()
        arq = arquivo_var.get().strip()
        _salvar_pasta_clone(path)
        di, df = inicio_var.get().strip(), fim_var.get().strip()

        if not path or not arq:
            messagebox.showwarning(
                _APP_TITLE,
                "Informe a pasta do clone e o arquivo.",
                parent=root,
            )
            return

        if not di and not df:
            since_dt, until_dt = _intervalo_ultimos_dias(60)
        elif not di or not df:
            messagebox.showwarning(
                _APP_TITLE,
                "Preencha “De” e “até” com mm/aaaa, ou apague os dois para usar os últimos 60 dias.",
                parent=root,
            )
            return
        else:
            dk = _normaliza_mm_yyyy_digitado(di)
            fk = _normaliza_mm_yyyy_digitado(df)
            if dk is None or fk is None:
                messagebox.showerror(
                    _APP_TITLE,
                    "Formato inválido. Use mm/aaaa em “De” e “até” (ex.: 3/2026).",
                    parent=root,
                )
                return
            if dk == _de_padrao_key and fk == _ate_padrao_key:
                since_dt, until_dt = _intervalo_ultimos_dias(60)
            else:
                try:
                    since_dt, until_dt = parse_intervalo_mes_ano(di, df)
                except ValueError as e:
                    messagebox.showerror(_APP_TITLE, str(e), parent=root)
                    return

        set_busy(True)
        status_var.set("Buscando…")

        do_fetch = fetch_var.get()

        def worker():
            err = None
            res = []
            try:
                res = buscar_branches_com_alteracao(
                    path,
                    arq,
                    since=since_dt,
                    until=until_dt,
                    refs_integracao=integracao_var.get().strip(),
                    fazer_fetch=do_fetch,
                    progress_callback=on_progress,
                )
            except Exception as e:
                err = e

            def finish():
                set_busy(False)
                if err:
                    messagebox.showerror(_APP_TITLE, str(err), parent=root)
                    status_var.set("Erro na busca.")
                    return
                if do_fetch:
                    _marcar_fetch_ok(path)
                    _marcar_auth_ok(path)
                    fetch_var.set(False)
                _preencher_tree(res)
                n = len(res)
                status_var.set(
                    f"Concluído: {n} branch(es) encontrada(s)."
                    if n
                    else "Concluído: nenhuma branch no período."
                )

            root.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    btn_run.configure(command=executar)

    root.after(500, verificar_git_ao_abrir)

    return root


if __name__ == "__main__":
    app = _criar_app()
    app.mainloop()
